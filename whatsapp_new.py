# Note: For proper working of this Script Good and Uninterepted Internet Connection is Required
# Keep all contacts unique
# Can save contact with their phone Number

# Import required packages
from selenium import webdriver
import phonenumbers
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
import datetime, time, os
import openpyxl as excel
import urllib.parse
from messages import get_all_message_templates

# function to read contacts from a text file

print("Fetching contacts details from sheet")
print("..................")

time.sleep(5)

def readSheet(fileName):
    lst = []
    file = excel.load_workbook(fileName)
    sheet = file.worksheets[0]
    bypass_title = False

    columns_name = [columnname.value.lower().replace(" ", "_") for columnname in next(sheet.iter_rows())]

    for row in sheet.iter_rows():
        if not bypass_title:
            bypass_title = True
            continue
        row_values = [row_item.value for row_item in row]
        lst.append(dict(zip(columns_name, row_values)))
        
    return lst

print("Validating Contact details............")

sheet_data = readSheet("contacts.xlsx")


def validateContacts():
    return sheet_data

# Not tested on Broadcast
contacts = validateContacts()

# Driver to open a browser
cService = webdriver.ChromeService(executable_path="/usr/bin/chromedriver")
cOptions = webdriver.ChromeOptions()
cOptions.add_argument("--user-data-dir=chrome-data")
driver = webdriver.Chrome(service= cService, options=cOptions)

attachment_image = "/home/sonu/Documents/lventer/whatsapp_senders/wbot/WhatsApp-bot-selenium/Lavanya Enterprises.png"


# #link to open a site
whatsapp_url = f"https://web.whatsapp.com"
driver.get(whatsapp_url)
time.sleep(60)

def get_message(**kwargs):
    all_templates = get_all_message_templates()
    if  kwargs.get("template_name"):
        message = all_templates[kwargs["template_name"]]
    else:
        message = all_templates["LAVANYA_ENGLISH"]
    return message.format(**kwargs)

message_not_sent = []

for contact in contacts:
    phone_number = contact["phone"]
    print(contact)
    try:
        time.sleep(20)
        # Find the search box and search for the contact
        search_box = driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="3"]')
        search_box.click()
        search_box.send_keys(phone_number)
        search_box.send_keys(Keys.RETURN)
        time.sleep(4)  # Wait for chat to open

        # Find the message input box and send the message
        print("Writing Message.........")
        message_box = driver.find_element(By.XPATH, '//div[@contenteditable="true"][@data-tab="10"]')
        message_box.click()
        # message_box.send_keys(LAVANYA_ENGLISH)

        # Splitting the message by lines and sending each line with Shift+Enter for a new line
        for line in get_message(**contact).splitlines():
            message_box.send_keys(line)
            message_box.send_keys(Keys.SHIFT + Keys.ENTER)
        message_box.send_keys(Keys.RETURN)  # Send the message

        time.sleep(4)  # Wait a bit for the message to be sent

        if attachment_image and os.path.exists(attachment_image):
            attach_button = driver.find_element(By.XPATH, '//div[@title="Attach"]')
            attach_button.click()
            time.sleep(3)  # Wait for the attach menu to open
            photo_input = driver.find_element(By.XPATH, '//input[@accept="image/*,video/mp4,video/3gpp,video/quicktime"]')
            photo_input.send_keys(attachment_image)
            time.sleep(5)  # Wait for the photo to upload
        # Send message
        send_button = driver.find_element(By.XPATH, '//span[@data-icon="send"]')
        send_button.click()
        # Send message
        print("Message sent to the user")
    except Exception as e:
            message_not_sent.append(phone_number)
            print(f"An error occurred with {contact}: {e}")

time.sleep(20)

print("This is list of peoples to who messages could not sent")
print(message_not_sent)

driver.quit()